'''
openpyxl, re, copy libraries required for execution. 
Script must be in same folder as working file
Takes a value against key and puts that value against where-ever key value is found in working sheet
'''
from openpyxl import load_workbook
from copy import copy
import openpyxl
import re

#returns first number in list
def return_num(lst):
    
    for value in lst:
        if value.isnumeric():
            return int(value)
    return 0

#### lookups each key in lookup sheet in designated cell, if found returns key value as per column specified
def lookup_2(lookup_value,lookup_sheet,lookup_column,default_value):

    for row in lookup_sheet.iter_rows():
        if row[0].value in lookup_value:
            return row[lookup_column].value
            break
    return default_value

#### lookups specified value in lookup sheet, if found returns key value as per column specified
def lookup(lookup_value,lookup_sheet,lookup_column,key_column):

    for row in lookup_sheet.iter_rows():
        if row[key_column].value==lookup_value:
            return row[lookup_column].value
            break
    return ''

##### Main Entry Function
def enter_values(working,lookup_workbook):
    
    fix_values=lookup_workbook['Fix Values']
    stone_color=lookup_workbook['Stone Color']
    gem_table=lookup_workbook['Gem Table']
    metal_weight_14k=lookup_workbook['Metal Weight 14k']
    metal_weight_18k=lookup_workbook['Metal Weight 18k']
    
    for index,row in enumerate(working.iter_rows()):
        

        if '-' in row[1].value:
            sep='-'
        elif '/' in row[1].value:
            sep='/'
        else:
            sep='|'
            
        sku_att=row[1].value.split(sep=sep) 
        
        if index==0: #Ignore headers
            continue
                
        elif row[7].value=='Metal Weight':
            
            lookup_value=return_num(sku_att)
            if '18K'.lower() in row[3].value.lower():
                row[8].value=lookup(lookup_value,metal_weight_18k,2,1)
            
            elif '14K'.lower() in row[3].value.lower():
                row[8].value=lookup(lookup_value,metal_weight_14k,1,0)
       
        
        elif row[7].value=='Metal Type':
            
            if 'White Gold'.lower() in row[3].value.lower():
                row[8].value='White Gold'
            elif 'Pink Gold'.lower() in row[3].value.lower() or 'Rose Gold'.lower() in row[3].value.lower():
                row[8].value='Rose Gold'
            elif 'Yellow Gold'.lower() in row[3].value.lower():
                row[8].value='Yellow Gold'
            else:
                row[8].value='Gold'
                
                
        elif row[7].value=='Metal Stamp':
            
            if '18k'.lower() in row[1].value.lower():
                row[8].value='18K'
            else:
                row[8].value='14K'
                
                
        elif row[7].value=='Chain Length Decimal Value':
            if 'Pendant'.lower() in row[3].value.lower():
                num_in_text = re.findall(r"[-+]?\d*\.\d+|\d+",row[3].value)
                count_18=num_in_text.count('18')
                count_18=num_in_text.count('16')
                
                if '18k' in row[3].value.lower():
                    if count_18>=2:
                        row[8].value=18
                    else:
                        row[8].value=16
                elif '14k' in row[3].value.lower():
                    if count_18>=1:
                        row[8].value=18
                    else:
                        row[8].value=16
                else:
                    row[8].value=""
            else:
                row[8].value=""
                
        elif row[7].value=='Stone Color':
            row[8].value=lookup_2(row[3].value,stone_color,1,'Clear')
        
        elif row[7].value=='Gem Type':
            row[8].value=lookup_2(row[3].value,gem_table,1,'Cubic Zirconia')
        
        else:
            #import fixed values
            row[8].value=lookup(row[7].value,fix_values,1,0)
            
    return working,lookup_workbook

### main program
main_file='Modified.xlsx'
working_workbook = load_workbook(filename=main_file)
working=working_workbook['Recommendations']
lookup_file='python instruction.xlsx'
lookup_workbook=load_workbook(filename=lookup_file)

working,lookup_workbook=enter_values(working,lookup_workbook)

working_workbook.save(filename='Modified1.xlsx')